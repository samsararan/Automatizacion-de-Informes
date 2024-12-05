
import openpyxl 
from docxtpl import DocxTemplate
from num2words import num2words
from datetime import datetime

# Función para manejar errores en forma segura
def manejar_error(mensaje, e):
    print(f"Error: {mensaje}")
    print(f"Detalle: {str(e)}")

# Intentar cargar el archivo de Excel
try:
    wb = openpyxl.load_workbook("planilla.xlsx")
except FileNotFoundError as e:
    manejar_error("No se encontró el archivo Excel 'planilla.xlsx'. Verifique el nombre y la ubicación.", e)
    exit()

# Intentar acceder a las hojas de cálculo
try:
    pestaña1 = wb["Hoja1"]
    pestaña2 = wb["Hoja2"]
except KeyError as e:
    manejar_error("Una de las hojas de Excel ('Hoja1' o 'Hoja2') no existe en el archivo.", e)
    exit()

wb.active = pestaña1

# Funciones para cambiar mes, formato monetario y convertir números a letras
def cambiar_mes(numero):
    meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
             "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    try:
        return meses[numero - 1]
    except IndexError:
        return "mes inválido"

def moneda(numero):
    try:
        return "${:,.2f}".format(numero).replace(",", ";").replace(".", ",").replace(";", ".")
    except Exception as e:
        manejar_error("No se pudo formatear el número como moneda.", e)
        return "Valor inválido"

def numero_a_letras_con_centavos(numero, idioma='es'):
    try:
        parte_entera = int(numero)
        centavos = round((numero - parte_entera) * 100)
        palabras_parte_entera = num2words(parte_entera, lang=idioma)
        if centavos > 0:
            palabras_centavos = num2words(centavos, lang=idioma)
            return f"{palabras_parte_entera} con {palabras_centavos} centavos"
        return palabras_parte_entera
    except Exception as e:
        manejar_error("Error al convertir el número a letras.", e)
        return "Número inválido"

# Procesar datos de las hojas
try:
    lista = [row for row in pestaña1.iter_rows(min_row=2, max_row=pestaña1.max_row-1, values_only=True)]
    datos = [row for row in pestaña2.iter_rows(min_row=2, max_row=2, values_only=True)]
except Exception as e:
    manejar_error("Error al leer datos de las hojas del archivo Excel.", e)
    exit()

# Generar textos dinámicos para el documento
try:
    rulo = ""
    for sub_lista in lista:
        rulo += f" - En el mes de {cambiar_mes((sub_lista[1]).month)} {sub_lista[1].year} (certificado N° {sub_lista[0]}) un factor de redeterminación definitivo de {str(sub_lista[9]).replace('.', ',')}.\n"

    clausula1 = ""
    for sub_lista in lista:
        clausula1 += f"   - En el mes de {cambiar_mes((sub_lista[1]).month)} {sub_lista[1].year} (certificado N° {sub_lista[0]}) un factor definitivo {str(sub_lista[9]).replace('.', ',')} comparado con el provisorio de {moneda(sub_lista[12])}.\n"
except Exception as e:
    manejar_error("Error al generar los textos dinámicos para el documento.", e)
    exit()

# Procesar fechas y certificados
try:
    meses = [row[0] for row in pestaña1.iter_rows(min_row=2, max_row=pestaña1.max_row-1, max_col=2, min_col=2, values_only=True)]
    mes_min, mes_max = meses[0].month, meses[-1].month
    fecha_min, fecha_max = f"{cambiar_mes(mes_min)} de {meses[0].year}", f"{cambiar_mes(mes_max)} de {meses[-1].year}"
except Exception as e:
    manejar_error("Error al calcular las fechas iniciales y finales del periodo.", e)
    exit()

# Generar los documentos
while True:
    try:
        opcion = input("Elija una opción:\n1 - Doc. General\n2 - Doc. Vialidad\n3 - Informe Rectificación\n4 - Salir\n")
        if opcion == "4":
            print("\nGracias por usar el programa. ¡Adiós!")
            break
        elif opcion in ["1", "2", "3"]:
            plantilla = {
                "1": "maa995template.docx",
                "2": "maa995template_v.docx",
                "3": "inf_rect.docx"
            }.get(opcion)

            documento = DocxTemplate(plantilla)
            for dato in datos:
                documento.render({
                    "empresa": dato[0],
                    "obra": dato[1],
                    "licitacion": dato[2],
                    "fecha_contr": dato[3],
                    "rulo": rulo,
                    "clausula1": clausula1,
                    "mes_min": fecha_min,
                    "mes_max": fecha_max,
                    "suma_dispos": moneda(12345),  # Ejemplo
                })

            documento.save(f"Documento_{opcion}.docx")
            print(f"Documento generado: Documento_{opcion}.docx")
        else:
            print("Por favor, seleccione una opción válida.")
    except Exception as e:
        manejar_error("Error durante la generación del documento.", e)


