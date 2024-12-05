import openpyxl 
from docxtpl import DocxTemplate
import docx
import locale
from datetime import datetime
from num2words import num2words
      
# Función para manejar errores en forma segura
def manejar_error(mensaje, e):
    print(f"Error: {mensaje}")
    print(f"Detalle: {str(e)}")

# Abrimos la planilla de excel

wb = openpyxl.load_workbook("planilla.xlsx")

# definimos las variables de cada hoja

pestaña2 = wb["Hoja2"]

pestaña1 = wb["Hoja1"]

#establecemos que la pestaña activa de la planilla es la 1

wb.active = pestaña1

# Creamos la funcion que nos va a cambiar las fechas por numero de meses a meses en letras mas adelante

def cambiar_mes(numero):
    meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
             "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    try:
        return meses[numero - 1]
    except IndexError:
        return "mes inválido"
    
def cambiar_mes(numero):
      if numero == 1:
            return "enero"
      if numero == 2:
            return "febrero"
      if numero == 3:
            return "marzo"
      if numero == 4:
            return "abril"
      if numero == 5:
            return "mayo"
      if numero == 6:
            return "junio"
      if numero == 7:
            return "julio"
      if numero == 8:
            return "agosto"
      if numero == 9:
            return "septiembre"
      if numero == 10:
            return "octubre"
      if numero == 11:
            return "noviembre"
      if numero == 12:
            return "diciembre"

# Acá creamos la lista con los valores de la clausula 1 del ACTA ACUERDO

lista = []

for row in pestaña1.iter_rows(min_row=2, max_row=pestaña1.max_row-1, values_only=True):
    lista.append(row)

print(lista)
# ACA CREAMOS EL PARRAFO DEL INFORME del punto 8

rulo = "" 

for sub_lista in lista:
    # for dato in sub_lista:
    rulo += f" - En el mes de {cambiar_mes((sub_lista[1]).month)} {sub_lista[1].year} (certificado N° {sub_lista[0]}) un factor de redeterminación definitivo de {str(sub_lista[9]).replace('.' , ',')} (resultante del 100% de la variación de referencia)\n"

punto_5 = ""


# ACA CREAMOS EL PARRAFO DEL ACTA ACUERDO EN BASE A LOS DATOS DE LA LISTA DE LA LINEA 57

clausula1 = ""

def moneda(numero):
    mon = "${:,.2f}".format(numero).replace(",",";").replace(".",",").replace(";",".")
    return mon

for sub_lista in lista:
      clausula1 += f"   - En el mes de {cambiar_mes((sub_lista[1]).month)} {sub_lista[1].year} (certificado N° {sub_lista[0]}) un factor de redeterminación definitivo {str(sub_lista[9]).replace('.' , ',')} (resultante del 100% de la variación de referencia), que comparado con el Factor de Redeterminación Provisorio calculado por el área comitente de {str(sub_lista[3]).replace('.',',')} y {str(sub_lista[4]).replace('.',',')} (resultante del 95% de la variación de referencia), da una diferencia a reconocer de {moneda(sub_lista[12])}.\n"

print(f" ****CLAUSULA**** \n{clausula1}")

print("+++LISTA+++ ")
print(lista)

wb.active = pestaña2

# ACA CREAMOS LA LISTA CON LOS DATOS PERTENECIENTES A LA OBRA, NOMBRE EMPRESA, LICITACION, FECHA, EXPEDIENTE, ETC

datos = []

for row in pestaña2.iter_rows(min_row=2, max_row=2, values_only=True):
    datos.append(row)

print(f" estos son los datos {datos}")


# ACA SACAMOS LOS MESES PUNTA DEL PERIODO

meses = []

for row in pestaña1.iter_rows(max_col=2, min_col=2, min_row=2, max_row=pestaña1.max_row-1, values_only=True):
     meses += row

print("meses")

print(meses)

mes_min = meses[0].month 

year_min = meses[0].year

mes_max = meses[-1].month

year_max = meses[-1].year

fecha_min =  f"{cambiar_mes(mes_min)} de {year_min}"

fecha_max = f"{cambiar_mes(mes_max)} de {year_max}"

# Aca sacamos los certificados punta de todo el periodo

num_cert = []

for row in pestaña1.iter_rows(max_col=1, min_col=1, min_row=2, max_row=pestaña1.max_row-1, values_only=True):
     num_cert += row

num_cert_min = num_cert[0]

num_cert_max = num_cert[-1]

# Aca sacamos los valores totales de los informes

saldos = []

for row in pestaña1.iter_cols(max_col=13, min_col=9, min_row= pestaña1.max_row ,max_row=pestaña1.max_row, values_only=True):
     saldos += row

print(f" estos son los saldos {saldos}")

# NUMEROS A LETRAS

def numero_a_letras_con_centavos(numero, idioma='es'):
    """
    Convierte un número en palabras, incluyendo los centavos si es necesario.
    
    :param numero: Número con decimales.
    :param idioma: Idioma para la conversión (por defecto 'es' - español).
    :return: Representación en palabras del número, incluyendo los centavos.
    """
    try:
        # Separar la parte entera y los centavos
        parte_entera = int(numero)
        centavos = round((numero - parte_entera) * 100)
        
        # Convertir la parte entera
        palabras_parte_entera = num2words(parte_entera, lang=idioma)
        
        # Convertir los centavos, si existen
        if centavos > 0:
            palabras_centavos = num2words(centavos, lang=idioma)
            resultado = f"{palabras_parte_entera} con {palabras_centavos} centavos"
        else:
            resultado = palabras_parte_entera
        
        return resultado
    except Exception as e:
        return f"Error al convertir: {e}"

suma_dispos_letras =  numero_a_letras_con_centavos(saldos[0], idioma="es")

recon_contra_letras = numero_a_letras_con_centavos(saldos[3], idioma="es")

dif_favor_letras = numero_a_letras_con_centavos(saldos[4], idioma="es")



print("DATOS")
print(datos)

while True:
      
      pregunta = input("Elija una opción:\n 1 - Doc. General \n 2 - Doc. de Vialidad \n 3 - Informe de recitificacion \n 4 - Salir \n")

      if pregunta == str(1):
            
            documento = DocxTemplate("maa995template.docx")
            for dato in datos:
                  documento.render({"empresa":dato[0],
                              "obra":dato[1],
                              "licitacion":dato[2],
                              "fecha_contr":dato[3],
                              "rulo":rulo,
                              "expediente":dato[4],
                              "clausula1":clausula1,
                              "mes_min": fecha_min,
                              "mes_max":fecha_max,
                              "num_cert_min":num_cert_min,
                              "num_cert_max":num_cert_max,
                              "suma_dispos":moneda(saldos[0]),
                              "recon_contra":moneda(saldos[3]),
                              "dif_favor":moneda(saldos[4]),
                              "suma_dispos_letras":suma_dispos_letras,
                              "recon_contra_letras":recon_contra_letras,
                              "dif_favor_letras":dif_favor_letras,
                              "localidad":dato[5]})

            documento.save("AA" + ".docx")

            # INFORME DEFINITIVAS 995 

            documento = DocxTemplate("infdef995template.docx")
            for dato in datos:
                  documento.render({"empresa":dato[0],
                              "obra":dato[1],
                              "licitacion":dato[2],
                              "fecha_contr":dato[3],
                              "rulo":rulo,
                              "expediente":dato[4],
                              "clausula1":clausula1,
                              "mes_min": fecha_min,
                              "mes_max":fecha_max,
                              "num_cert_min":num_cert_min,
                              "num_cert_max":num_cert_max,
                              "suma_dispos":moneda(saldos[0]),
                              "recon_contra":moneda(saldos[3]),
                              "dif_favor":moneda(saldos[4]),
                              "suma_dispos_letras":suma_dispos_letras,
                              "recon_contra_letras":recon_contra_letras,
                              "dif_favor_letras":dif_favor_letras,
                              "localidad":dato[5]})

            documento.save("INF-DEF " +".docx")

            print("Se generaron los archivos de areas generales") 
            break

      # VIALIDAD

      if pregunta == str(2):
            
            documento = DocxTemplate("maa995template_v.docx")
            for dato in datos:
                  documento.render({"empresa":dato[0],
                              "obra":dato[1],
                              "licitacion":dato[2],
                              "fecha_contr":dato[3],
                              "rulo":rulo,
                              "expediente":dato[4],
                              "clausula1":clausula1,
                              "mes_min": fecha_min,
                              "mes_max":fecha_max,
                              "num_cert_min":num_cert_min,
                              "num_cert_max":num_cert_max,
                              "suma_dispos":moneda(saldos[0]),
                              "recon_contra":moneda(saldos[3]),
                              "dif_favor":moneda(saldos[4]),
                              "suma_dispos_letras":suma_dispos_letras,
                              "recon_contra_letras":recon_contra_letras,
                              "dif_favor_letras":dif_favor_letras})

            documento.save("AA_V" + ".docx")

            documento = DocxTemplate("infdef995template_v.docx")
            for dato in datos:
                  documento.render({"empresa":dato[0],
                              "obra":dato[1],
                              "licitacion":dato[2],
                              "fecha_contr":dato[3],
                              "rulo":rulo,
                              "expediente":dato[4],
                              "clausula1":clausula1,
                              "mes_min": fecha_min,
                              "mes_max":fecha_max,
                              "num_cert_min":num_cert_min,
                              "num_cert_max":num_cert_max,
                              "suma_dispos":moneda(saldos[0]),
                              "recon_contra":moneda(saldos[3]),
                              "dif_favor":moneda(saldos[4]),
                              "suma_dispos_letras":suma_dispos_letras,
                              "recon_contra_letras":recon_contra_letras,
                              "dif_favor_letras":dif_favor_letras})

            documento.save("INF-DEF_V " +".docx")

            print("Se generaron los informes de vialdad")
            break

      # INFORME DE RECTIFICACION

      if pregunta == "3":

            documento = DocxTemplate("inf_rect.docx")
            for dato in datos:
                  documento.render({"empresa":dato[0],
                              "obra":dato[1],
                              "licitacion":dato[2],
                              "fecha_contr":dato[3],
                              "rulo":rulo,
                              "expediente":dato[4],
                              "clausula1":clausula1,
                              "mes_min": fecha_min,
                              "mes_max":fecha_max,
                              "num_cert_min":num_cert_min,
                              "num_cert_max":num_cert_max,
                              "suma_dispos":moneda(saldos[0]),
                              "recon_contra":moneda(saldos[3]),
                              "dif_favor":moneda(saldos[4]),
                              "suma_dispos_letras":suma_dispos_letras,
                              "recon_contra_letras":recon_contra_letras,
                              "dif_favor_letras":dif_favor_letras})

            documento.save("INFORME DE RECTIFICACION" +".docx")

            print("Se generó el archivo de informes de rectificación")
            break

      if pregunta == "4":
            print("\n *** Gracias por usar GOYO-SOFT *** \n")
            break
      else: 
            print("*** Ingrese una opción válida *** \n")
            continue

input("Presiona Enter para salir...")